#!/usr/bin/env python3
# -*- coding: UTF-8 -*-
import re
import traceback
from io import BytesIO

from flask import Blueprint, request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import IllegalCharacterError

from app import db
from app.models import WorkOrder, Asset, Api, scanList
from datetime import datetime

from app.source.commontParams import json_response

common_bp = Blueprint('common', __name__)
allow_file_type = ['xlsx', 'xls']


@common_bp.route('/api/common/export', methods=["GET"])
def commonExport():
    common_type = request.args.get("type")
    if 'risk' == common_type:
        try:
            result = WorkOrder.query.all()
            if not result:
                return "No data to export", 200
            # 2. 创建 Excel 并写入内存
            wb = Workbook()
            ws = wb.active
            ws.title = "工单数据"

            # 写入表头
            ws.append(["ID", "风险名称", "风险类型", "风险等级", "部门", "业务线", "当前处理人", "状态", "创建时间"])
            # 写入数据
            for wOrder in result:
                ws.append([wOrder.id, wOrder.riskName, wOrder.riskType, wOrder.riskLevel, wOrder.department, wOrder.businessline, wOrder.currentOwner, wOrder.status, wOrder.createAt])

            # 3. 将 Excel 数据写入内存字节流
            output = BytesIO()  # 创建内存流
            wb.save(output)  # 保存 Excel 到内存流
            output.seek(0)  # 重置文件指针到开头（否则读取不到内容）

            now = datetime.now()
            timestamp = now.timestamp()
            # 4. 直接返回内存流
            return send_file(
                output,
                as_attachment=True,
                download_name="工单数据"+str(timestamp)+".xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            traceback.print_exc()
            return f"export fail: {str(e)}", 500
    elif 'asset' == common_type:
        try:
            result = Asset.query.all()
            if not result:
                return "No data to export", 200
            # 2. 创建 Excel 并写入内存
            wb = Workbook()
            ws = wb.active
            ws.title = "应用资产数据"

            # 写入表头
            ws.append(["ID", "系统名称", "url", "内外网", "关联接口数", "部门", "业务线", "应用负责人", "技术栈", "创建时间"])
            # 写入数据
            for asset in result:
                ws.append([asset.id, asset.assetName, asset.url, asset.insideOroutside, asset.linkedApiNum, asset.department, asset.businessline, asset.owner, asset.technologyStack, asset.createAt])

            # 3. 将 Excel 数据写入内存字节流
            output = BytesIO()  # 创建内存流
            wb.save(output)  # 保存 Excel 到内存流
            output.seek(0)  # 重置文件指针到开头（否则读取不到内容）

            now = datetime.now()
            timestamp = now.timestamp()
            # 4. 直接返回内存流
            return send_file(
                output,
                as_attachment=True,
                download_name="应用资产数据"+str(timestamp)+".xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            traceback.print_exc()
            return f"export fail: {str(e)}", 500
    elif 'api' == common_type:
        try:
            result = Api.query.all()
            if not result:
                return "No data to export", 200
            # 2. 创建 Excel 并写入内存
            wb = Workbook()
            ws = wb.active
            ws.title = "接口资产数据"

            # 写入表头
            ws.append(["ID", "接口路径", "描述", "部门", "业务线", "请求", "响应", "创建时间", "上一次更新时间"])
            # 写入数据
            for api in result:
                regex = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]')
                req = regex.sub('', api.request)
                resp = regex.sub('', api.response)
                ws.append([api.id, api.apiPath, api.describe, api.department, api.businessline, req, resp, api.createAt, api.updateAt])

            # 3. 将 Excel 数据写入内存字节流
            output = BytesIO()  # 创建内存流
            wb.save(output)  # 保存 Excel 到内存流
            output.seek(0)  # 重置文件指针到开头（否则读取不到内容）

            now = datetime.now()
            timestamp = now.timestamp()
            # 4. 直接返回内存流
            return send_file(
                output,
                as_attachment=True,
                download_name="接口资产数据"+str(timestamp)+".xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            traceback.print_exc()
            return f"export fail: {str(e)}", 500
    elif 'scanList' == common_type:
        try:
            result = scanList.query.all()
            if not result:
                return "No data to export", 200
            # 2. 创建 Excel 并写入内存
            wb = Workbook()
            ws = wb.active
            ws.title = "扫描接口数据"

            # 写入表头
            ws.append(["ID", "扫描路径", "相似度", "扫描时间"])
            # 写入数据
            for sList in result:
                ws.append([sList.id, sList.scanPath, sList.similarity, sList.scanAt])

            # 3. 将 Excel 数据写入内存字节流
            output = BytesIO()  # 创建内存流
            wb.save(output)  # 保存 Excel 到内存流
            output.seek(0)  # 重置文件指针到开头（否则读取不到内容）

            now = datetime.now()
            timestamp = now.timestamp()
            # 4. 直接返回内存流
            return send_file(
                output,
                as_attachment=True,
                download_name="扫描接口数据"+str(timestamp)+".xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            traceback.print_exc()
            return f"export fail: {str(e)}", 500
    return "type is error", 200


@common_bp.route('/api/common/upload', methods=['POST'])
def commonUpload():
    common_type = request.args.get("type")
    if 'asset' == common_type:
        if 'file' not in request.files:
            return json_response(data=[], count=0, msg='没有文件部分', status=0)
        file = request.files['file']
        if file.filename == '':
            return json_response(data=[], count=0, msg='没有选择文件', status=0)
        if file and '.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allow_file_type:
            file_stream = BytesIO(file.read())
            print(file_stream)
            success, error = parse_excel_in_memory(file_stream)
            if success:
                return json_response(data=[], count=0, msg="添加资产成功", status=200)
            else:
                return json_response(data=[], count=0, msg="添加资产失败，url可能已经存在", status=400)
        return json_response(data=[], count=0, msg="不支持文件类型", status=400)


def clean_illegal_chars(value):
    """清理Excel中的非法字符"""
    if value is None:
        return ""

    # 转换为字符串
    if not isinstance(value, str):
        value = str(value)

    # 移除ASCII控制字符（保留制表符\t）
    regex = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]')
    cleaned_value = regex.sub('', value)

    return cleaned_value


def parse_excel_in_memory(file_stream):
    """直接在内存中解析Excel文件并返回数据列表"""
    try:
        # 直接从文件流加载工作簿，不保存到磁盘
        workbook = load_workbook(file_stream, data_only=True)
        sheet = workbook.active  # 获取第一个工作表
        # 假设第一行是标题行，从第二行开始读取数据
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # 清理每行数据中的非法字符
            cleaned_row = [clean_illegal_chars(cell) for cell in row]

            # 根据你的Excel列数和顺序调整
            row_data = {
                'name': cleaned_row[0] if len(cleaned_row) > 0 else None,
                'email': cleaned_row[1] if len(cleaned_row) > 1 else None,
                'phone': cleaned_row[2] if len(cleaned_row) > 2 else None,
                'department': cleaned_row[3] if len(cleaned_row) > 3 else None,
                'join_date': cleaned_row[4] if len(cleaned_row) > 4 else None
            }
            asset = Asset()
            asset.assetName = cleaned_row[0] if len(cleaned_row) > 0 else ""
            asset.url = cleaned_row[1] if len(cleaned_row) > 0 else ""
            asset.insideOroutside = cleaned_row[2] if len(cleaned_row) > 0 else ""
            asset.linkedApiNum = cleaned_row[3] if len(cleaned_row) > 0 else 0
            asset.department = cleaned_row[4] if len(cleaned_row) > 0 else 0
            asset.businessline = cleaned_row[5] if len(cleaned_row) > 0 else 0
            asset.owner = cleaned_row[6] if len(cleaned_row) > 0 else 0
            asset.technologyStack = cleaned_row[7] if len(cleaned_row) > 0 else 0
            # 获取当前时间
            now = datetime.now()
            # 格式化为字符串
            formatted_time = now.strftime('%Y-%m-%d %H:%M:%S')
            asset.createAt = formatted_time
            db.session.add(asset)
            db.session.commit()
    except IllegalCharacterError as e:
        return False, f"Excel包含非法字符: {str(e)}"
    except Exception as e:
        traceback.print_exc()
        return False, f"添加资产出错： {str(e)}"
